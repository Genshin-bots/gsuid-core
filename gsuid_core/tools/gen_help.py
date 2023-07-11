import copy
import json
from pathlib import Path

from openpyxl import load_workbook

sample = {
    'name': '',
    'desc': '',
    'eg': '',
    'need_ck': False,
    'need_sk': False,
    'need_admin': False,
}

result = {}

HELP_PATH = Path(__file__).parent / 'Help.xlsx'
OUTPUT_PATH = Path(__file__).parent / 'Help.json'

wb = load_workbook(str(HELP_PATH))
ws = wb.active

module_name_str = ''
for row in range(2, 999):
    # 跳过空白行
    if not ws.cell(row, 2).value:
        continue

    _sample = copy.deepcopy(sample)

    # 将第一列读取为模块名
    if ws.cell(row, 1):
        if ws.cell(row, 1).value is not None:
            module_name_str = ws.cell(row, 1).value

    # if module_name_str is None and not isinstance(module_name_str, str):
    #    continue

    # 第二列为功能名
    _sample['name'] = ws.cell(row, 2).value
    # 第三列为详细信息
    _sample['desc'] = ws.cell(row, 3).value
    # 第四列为使用例
    _sample['eg'] = ws.cell(row, 4).value

    if ws.cell(row, 5).value == '是':
        _sample['need_ck'] = True

    if ws.cell(row, 6).value == '是':
        _sample['need_sk'] = True

    if ws.cell(row, 7).value == '是':
        _sample['need_admin'] = True

    if isinstance(module_name_str, str):
        module_name = module_name_str.split(' | ')[0]
        module_desc = module_name_str.split(' | ')[1]
        if module_name not in result:
            result[module_name] = {'desc': module_desc, 'data': []}

        result[module_name]['data'].append(_sample)

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)
